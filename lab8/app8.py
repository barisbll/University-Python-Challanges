import argparse
import csv
import openpyxl
from openpyxl.styles import Font, Color, Alignment
from openpyxl.styles import colors


#Main function
def run():

    # list is the list of arguments
    list = read_stdin()
    dataset_file_str = list[0]
    excell_name = str(list[1])[2:-2]

    # Put data in correct data structure
    data_dict = read_csv(dataset_file_str)

    # If -o option used in stdin use excell_operations
    if list[1] == None:
        operations(data_dict)

    else:
        excell_operations(data_dict ,excell_name)


# Read data from stdin returns list of arguments used
def read_stdin():

    parser = argparse.ArgumentParser()
    parser.add_argument('dataset', help='A csv file to be read', action='store')
    parser.add_argument('-o', help='Allow user to save the operations in given name', action='store', required=False, nargs=1)
    parser.add_argument('-i', help='help option', action='help')

    args = parser.parse_args()


    try:
        if vars(args)['dataset'] == None:
            raise FileNotFoundError

        if args.dataset.split('.')[-1] != 'csv':
            raise FileNotCSVError(vars(args)['dataset'].split('.')[-1])

        if 'o' in vars(args):
            return [vars(args)['dataset'], vars(args)['o']]


        return [vars(args)['dataset']]


    except FileNotFoundError as fnf:
        print(fnf)

    except FileNotCSVError as fncsv:
        print(fncsv)


# If dataset's extention isn't csv return this error
class FileNotCSVError(Exception):

    def __init__(self, extentionType, message="Error extention is not CSV"):
        self.message = f'{message} it is {extentionType} '
        super().__init__(self.message)


# Takes filename as str, reads and puts into a dictionary
# Returns dictionary
def read_csv(file_str):

    dict = {
        'gender': [],
        'race/ethnicity' : [],
        'parental level of education': [],
        'lunch': [],
        'test preparation course': [],
        'math score': [],
        'reading score': [],
        'writing score': [],
    }

    with open(file_str) as data_file:
        dict_reader = csv.DictReader(data_file)
        for row in dict_reader:
            dict['gender'].append(row['gender'])
            dict['race/ethnicity'].append(row['race/ethnicity'])
            dict['parental level of education'].append(row['parental level of education'])
            dict['lunch'].append(row['lunch'])
            dict['test preparation course'].append(row['test preparation course'])
            dict['math score'].append(row['math score'])
            dict['reading score'].append(row['reading score'])
            dict['writing score'].append(row['writing score'])

    return dict


# Makes some operations prints to the stdout
def operations(data_dict):

    # Aggregation Example
    print('Percentage of childrens gender')

    male_ctr = 0
    female_ctr = 0

    for e in data_dict['gender']:
        if e == 'male':
            male_ctr += 1
        else:
            female_ctr += 1

    print(f'There are {male_ctr} males and {female_ctr} females ')
    print(f'Female percentage is {female_ctr / (female_ctr + male_ctr)} \n')

    # Statistical Example
    sum_maths = 0
    sum_reading = 0
    sum_writing = 0
    total_length = len(data_dict['math score'])

    for e in data_dict['math score']:
        sum_maths += int(e)

    for e in data_dict['reading score']:
        sum_reading += int(e)

    for e in data_dict['writing score']:
        sum_writing += int(e)

    print('Average grades for specified classes')
    print(f'Average grade for mathematics {sum_maths/total_length}')
    print(f'Average grade for reading {sum_reading / total_length}')
    print(f'Average grade for writing {sum_writing / total_length} \n')

    # Summary Example

    print(f'Total length of the dataset {total_length}')
    print(f'Total number of classes {len(set(data_dict["race/ethnicity"]))}')
    print(f'These classes are {set(data_dict["race/ethnicity"])}')


# Makes some operations writes to excell and saves.
def excell_operations(data_dict, excell_name):

    # Aggregation Example
    male_ctr = 0
    female_ctr = 0

    for e in data_dict['gender']:
        if e == 'male':
            male_ctr += 1
        else:
            female_ctr += 1

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title= 'Percentage of Genders'

    # Styling Sheet 1

    a1 = sheet['A1']
    a1.font = Font(italic=True, bold=True, color='2F6B59')
    a1.alignment = Alignment(wrap_text=True)

    b1 = sheet['B1']
    b1.font = Font(italic=True, bold=True, color='2F6B59')
    b1.alignment = Alignment(wrap_text=True)

    c1 = sheet['C1']
    c1.font = Font(italic=True, bold=True, color='2F6B59')
    c1.alignment = Alignment(wrap_text=True)

    d1 = sheet['D1']
    d1.font = Font(italic=True, bold=True, color='2F6B59')
    d1.alignment = Alignment(wrap_text=True)



    sheet.cell(row=1,column=1, value='Genders')

    for r in range(2, len(data_dict['gender']) + 2):
        sheet.cell(row=r,column=1,value=data_dict['gender'][r-2])
        tempRow = sheet.cell(row=r, column=1)
        if tempRow.value == 'female':
            tempRow.font = Font(color='BC4959')
        else:
            tempRow.font = Font(color='9213D5')

    sheet.cell(row=1, column=2, value='Female Total Number')
    sheet.cell(row=2, column=2, value=female_ctr)

    sheet.cell(row=1, column=3, value='Male Total Number')
    sheet.cell(row=2, column=3, value=male_ctr)

    sheet.cell(row=1, column=4, value='Female Percentage')
    sheet.cell(row=2, column=4, value=female_ctr / (female_ctr + male_ctr))


    # Statistical Example
    sum_maths = 0
    sum_reading = 0
    sum_writing = 0
    total_length = len(data_dict['math score'])

    for e in data_dict['math score']:
        sum_maths += int(e)

    for e in data_dict['reading score']:
        sum_reading += int(e)

    for e in data_dict['writing score']:
        sum_writing += int(e)

    sheet2 = workbook.create_sheet('Average For Classes')
    sheet2.cell(row=1, column=1, value='Mathematics Scores')
    sheet2.cell(row=1, column=2, value='Reading Scores')
    sheet2.cell(row=1, column=3, value='Writing Scores')

    for r in range(2, len(data_dict['reading score']) + 2):
        temp1 = sheet2.cell(row=r, column=1, value=int(data_dict['math score'][r-2]))
        temp2 = sheet2.cell(row=r, column=2, value=int(data_dict['reading score'][r-2]))
        temp3 = sheet2.cell(row=r, column=3, value=int(data_dict['writing score'][r-2]))

        temp1.font = Font(color='BC4959')
        temp2.font = Font(color='9213D5')
        temp3.font = Font(color='92ECFF')


    sheet2.cell(row=1,column=4,value='Average of Mathematics')
    sheet2.cell(row=1, column=5, value='Average of Reading')
    sheet2.cell(row=1, column=6, value='Average of Writing')

    sheet2.cell(row=2, column=4, value=sum_maths/total_length)
    sheet2.cell(row=2, column=5, value=sum_reading / total_length)
    sheet2.cell(row=2, column=6, value=sum_writing / total_length)

    a1 = sheet2.cell(row=1, column=1)
    a1.font = Font(italic=True, bold=True, color='2F6B59')
    a1.alignment = Alignment(wrap_text=True)

    b1 = sheet2.cell(row=1, column=2)
    b1.font = Font(italic=True, bold=True, color='2F6B59')
    b1.alignment = Alignment(wrap_text=True)

    c1 = sheet2.cell(row=1, column=3)
    c1.font = Font(italic=True, bold=True, color='2F6B59')
    c1.alignment = Alignment(wrap_text=True)

    d1 = sheet2.cell(row=1, column=4)
    d1.font = Font(italic=True, bold=True, color='2F6B59')
    d1.alignment = Alignment(wrap_text=True)

    e1 = sheet2.cell(row=1, column=5)
    e1.font = Font(italic=True, bold=True, color='2F6B59')
    e1.alignment = Alignment(wrap_text=True)

    f1 = sheet2.cell(row=1, column=6)
    f1.font = Font(italic=True, bold=True, color='2F6B59')
    f1.alignment = Alignment(wrap_text=True)

    g1 = sheet2.cell(row=1, column=7)
    g1.font = Font(italic=True, bold=True, color='2F6B59')
    g1.alignment = Alignment(wrap_text=True)


    # Summary Example

    sheet3 = workbook.create_sheet('Total Number of Ethnique Groups')
    a1 = sheet3.cell(row=1, column=1, value='Ethnique Data')
    b1 = sheet3.cell(row=1, column=2, value='Unique Ethnic Groups')
    c1 = sheet3.cell(row=1, column=3, value='Total Number of Ethnicity Data')

    a1.font = Font(italic=True, bold=True, color='2F6B59')
    a1.alignment = Alignment(wrap_text=True)

    b1.font = Font(italic=True, bold=True, color='2F6B59')
    b1.alignment = Alignment(wrap_text=True)

    c1.font = Font(italic=True, bold=True, color='2F6B59')
    c1.alignment = Alignment(wrap_text=True)


    for r in range(2, len(set(data_dict["race/ethnicity"])) + 2):
        t = sheet3.cell(row=r, column=2, value=list(set(data_dict["race/ethnicity"]))[r-2])
        t.font = Font(color='BC4959')

    sheet3.cell(row=2, column=3, value=len(data_dict["race/ethnicity"]))

    for r in range(2, len(data_dict["race/ethnicity"]) + 2):
        t = sheet3.cell(row=r, column=1, value=data_dict["race/ethnicity"][r-2])
        t.font = Font(color='9213D5')


    workbook.save(f'{excell_name}.xlsx')


run()